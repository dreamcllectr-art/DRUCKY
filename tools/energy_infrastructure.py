"""Energy Infrastructure Intelligence — interconnection queues, tech cost curves, regulatory signals.

Three data pillars for deep energy/utilities sector analysis:

1. **Interconnection Queue Analysis** (CAISO direct + gridstatus)
   - PJM, MISO, CAISO, ERCOT, SPP, NYISO, ISO-NE
   - Queue depth by fuel type, MW capacity, completion rates, bottleneck detection
   - Quarterly refresh (queue data doesn't change daily)

2. **NREL Technology Cost Curves** (ATB S3/CSV downloads)
   - LCOE, capex, O&M for solar, wind, battery, nuclear, gas
   - Historical + projected through 2050
   - Annual refresh

3. **DOE/Congressional Activity** (Congress.gov API + DOE press scrape)
   - Energy bills in current Congress, committee status, cosponsor count
   - DOE LPO conditional commitments and closed loans
   - Weekly refresh

Architecture: Data -> SQLite cache -> sector expert context -> intelligence reports

Usage:
    python -m tools.energy_infrastructure --all          # full refresh
    python -m tools.energy_infrastructure --queues       # interconnection only
    python -m tools.energy_infrastructure --costs        # NREL costs only
    python -m tools.energy_infrastructure --regulatory   # Congress + DOE only
"""

import sys
import os
import json
import time
import argparse
from datetime import date, datetime
from pathlib import Path

_project_root = str(Path(__file__).parent.parent)
if _project_root not in sys.path:
    sys.path.insert(0, _project_root)

import requests

from tools.config import GEMINI_API_KEY, GEMINI_BASE, GEMINI_MODEL
from tools.db import init_db, query, get_conn

# ===================================================================
#  CONFIGURATION
# ===================================================================

# Congress.gov API
CONGRESS_API_KEY = os.getenv("CONGRESS_API_KEY", "")
CONGRESS_API_BASE = "https://api.congress.gov/v3"
CURRENT_CONGRESS = 119  # 2025-2026

# NREL ATB direct download (versioned path)
NREL_ATB_URL = "https://oedi-data-lake.s3.amazonaws.com/ATB/electricity/csv/2024/v3.0.0/ATBe.csv"

# Technology names as they appear in the ATB CSV 'technology' column
NREL_TECH_MAP = {
    "utilitypv": "Utility Solar",
    "landbasedwind": "Onshore Wind",
    "offshorewind": "Offshore Wind",
    "utility-scalepv-plus-battery": "Solar + Storage",
    "nuclear": "Nuclear",
    "naturalgas": "Natural Gas",
    "csp": "Concentrated Solar",
    "geothermal": "Geothermal",
    "hydropower": "Hydro",
    "biopower": "Biomass",
}

# Fuel type normalization for queue analysis
FUEL_NORMALIZE = {
    "solar": "Solar", "solar photovoltaic": "Solar", "photovoltaic": "Solar",
    "wind": "Wind", "onshore wind": "Wind", "offshore wind": "Offshore Wind",
    "battery": "Battery Storage", "battery storage": "Battery Storage",
    "energy storage": "Battery Storage",
    "natural gas": "Natural Gas", "gas": "Natural Gas", "ng": "Natural Gas",
    "nuclear": "Nuclear", "nuclear fission": "Nuclear",
    "hybrid": "Hybrid", "solar + storage": "Hybrid",
    "coal": "Coal",
    "pumped": "Hydro", "hydro": "Hydro", "hydroelectric": "Hydro",
    "geothermal": "Geothermal", "biomass": "Biomass",
}

# Tickers affected by queue/cost/regulatory data
QUEUE_AFFECTED_TICKERS = {
    "Solar": ["FSLR", "ENPH", "SEDG", "RUN", "NOVA", "ARRY"],
    "Wind": ["GE", "SHLS", "TPI"],
    "Battery Storage": ["FLUENCE", "STEM", "ENVX", "QS"],
    "Nuclear": ["CEG", "VST", "NRG", "SMR", "OKLO", "NNE"],
    "Natural Gas": ["EQT", "RRC", "AR", "SWN", "CTRA"],
    "Hybrid": ["NEE", "AES", "BEP"],
    "Utilities (Interconnection)": ["NEE", "DUK", "SO", "AEP", "D", "EXC", "SRE", "WEC", "ES", "ED"],
}


# ===================================================================
#  DATABASE SCHEMA
# ===================================================================

_SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS interconnection_queue (
    iso TEXT NOT NULL,
    project_name TEXT,
    developer TEXT,
    capacity_mw REAL,
    fuel_type TEXT,
    fuel_normalized TEXT,
    status TEXT,
    state TEXT,
    county TEXT,
    queue_date TEXT,
    expected_cod TEXT,
    interconnection_point TEXT,
    fetched_date TEXT DEFAULT (date('now')),
    UNIQUE(iso, project_name, capacity_mw)
);
CREATE INDEX IF NOT EXISTS idx_iq_iso_fuel ON interconnection_queue(iso, fuel_normalized);
CREATE INDEX IF NOT EXISTS idx_iq_status ON interconnection_queue(status);

CREATE TABLE IF NOT EXISTS interconnection_queue_summary (
    date TEXT NOT NULL,
    iso TEXT NOT NULL,
    fuel_type TEXT NOT NULL,
    active_count INTEGER,
    active_mw REAL,
    withdrawn_pct REAL,
    avg_queue_days REAL,
    median_cod_year INTEGER,
    PRIMARY KEY (date, iso, fuel_type)
);

CREATE TABLE IF NOT EXISTS nrel_cost_curves (
    technology TEXT NOT NULL,
    scenario TEXT NOT NULL,
    year INTEGER NOT NULL,
    metric TEXT NOT NULL,
    value REAL,
    unit TEXT,
    fetched_date TEXT DEFAULT (date('now')),
    PRIMARY KEY (technology, scenario, year, metric)
);

CREATE TABLE IF NOT EXISTS energy_legislation (
    bill_id TEXT PRIMARY KEY,
    congress INTEGER,
    bill_type TEXT,
    bill_number INTEGER,
    title TEXT,
    introduced_date TEXT,
    latest_action TEXT,
    latest_action_date TEXT,
    sponsor TEXT,
    cosponsor_count INTEGER,
    subjects TEXT,
    status TEXT,
    energy_relevance_score REAL,
    affected_sectors TEXT,
    affected_tickers TEXT,
    summary TEXT,
    fetched_date TEXT DEFAULT (date('now'))
);
CREATE INDEX IF NOT EXISTS idx_eleg_congress ON energy_legislation(congress, latest_action_date DESC);

CREATE TABLE IF NOT EXISTS energy_regulatory_signals (
    date TEXT NOT NULL,
    source TEXT NOT NULL,
    signal_type TEXT NOT NULL,
    headline TEXT,
    detail TEXT,
    affected_sectors TEXT,
    affected_tickers TEXT,
    impact_score REAL,
    PRIMARY KEY (date, source, headline)
);
"""


def _ensure_tables():
    """Create tables if they don't exist."""
    conn = get_conn()
    conn.executescript(_SCHEMA_SQL)
    conn.commit()


def _normalize_fuel(raw):
    """Normalize fuel type strings to standard categories."""
    if not raw:
        return "Unknown"
    raw_lower = str(raw).strip().lower()
    for key, val in FUEL_NORMALIZE.items():
        if key in raw_lower:
            return val
    return str(raw).strip().title()


# ===================================================================
#  1. INTERCONNECTION QUEUE ANALYSIS
# ===================================================================

def _fetch_caiso_queue(conn, today):
    """Fetch CAISO queue from their public Excel download."""
    try:
        import openpyxl
        import io
    except ImportError:
        print("      openpyxl not installed. Run: venv/bin/pip install openpyxl")
        return {"projects": 0, "mw": 0, "error": "openpyxl not installed"}

    try:
        url = "https://www.caiso.com/documents/publicqueuereport.xlsx"
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()

        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
        ws = wb.active

        # CAISO Excel has merged header rows; actual column names are on row 3
        # Find the header row by looking for "Project Name"
        header_row = 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            vals = [str(c or "").strip().lower() for c in row]
            if "project name" in vals:
                header_row = row_idx
                break

        headers = [str(c.value or "").strip().lower().replace("\n", " ")
                   for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not any(row):
                continue
            rd = dict(zip(headers, row))

            # CAISO uses "net mws to grid", "mw-1", "mw-2", "mw-3"
            capacity = None
            for col in ["net mws to grid", "mw-1", "net mw", "capacity (mw)", "mw"]:
                if col in rd and rd[col]:
                    try:
                        capacity = float(rd[col])
                    except (ValueError, TypeError):
                        pass
                    if capacity is not None:
                        break

            fuel = rd.get("fuel-1", rd.get("fuel type", rd.get("resource type", rd.get("fuel", ""))))
            status = rd.get("application status", rd.get("status", rd.get("queue status", "")))
            name = str(rd.get("project name", rd.get("name", "")))[:200]
            state = rd.get("state", rd.get("county", "CA"))
            county = rd.get("county", "")
            queue_dt = rd.get("queue date", "")
            cod = rd.get("current on-line date", rd.get("proposed on-line date (as filed with ir)", ""))
            intercon = rd.get("station or transmission line", "")

            # Normalize datetime objects to date strings
            def _to_date_str(val):
                if hasattr(val, "strftime"):
                    return val.strftime("%Y-%m-%d")
                if val:
                    s = str(val).strip()[:10]
                    return s if s and s != "None" else None
                return None
            queue_dt = _to_date_str(queue_dt)
            cod = _to_date_str(cod)

            rows.append((
                "CAISO", name, None, capacity,
                str(fuel or ""), _normalize_fuel(str(fuel or "")),
                str(status or ""), str(state or "")[:50],
                str(county or "")[:50] if county else None,
                queue_dt, cod,
                str(intercon or "")[:200] if intercon else None,
                today,
            ))
        wb.close()

        if rows:
            conn.executemany("""
                INSERT OR REPLACE INTO interconnection_queue
                (iso, project_name, developer, capacity_mw, fuel_type, fuel_normalized,
                 status, state, county, queue_date, expected_cod, interconnection_point, fetched_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, rows)
            conn.commit()

        active = [r for r in rows if "withdraw" not in (r[6] or "").lower()]
        mw = sum(r[3] or 0 for r in active)
        print(f"      CAISO: {len(active)} active projects, {mw:,.0f} MW")
        return {"projects": len(active), "mw": mw}
    except Exception as e:
        print(f"      CAISO FAILED: {e}")
        return {"projects": 0, "mw": 0, "error": str(e)}


def _fetch_gridstatus_queues(conn, today):
    """Fetch queue data via gridstatus library (if installed)."""
    try:
        import gridstatus
    except ImportError:
        return {}

    iso_map = {
        "PJM": gridstatus.PJM,
        "MISO": gridstatus.MISO,
        "ERCOT": gridstatus.Ercot,
        "SPP": gridstatus.SPP,
        "NYISO": gridstatus.NYISO,
        "ISONE": gridstatus.ISONE,
    }

    results = {}
    for iso_name, iso_class in iso_map.items():
        print(f"    Fetching {iso_name} queue via gridstatus...")
        try:
            iso = iso_class()
            df = iso.get_interconnection_queue()
            if df is None or df.empty:
                results[iso_name] = {"projects": 0, "mw": 0}
                continue

            cols = {c.lower().replace(" ", "_"): c for c in df.columns}
            rows = []
            for _, row in df.iterrows():
                capacity = None
                for col_name in ["capacity_mw", "capacity_(mw)", "mw",
                                 "summer_capacity_mw", "nameplate_capacity_mw"]:
                    if col_name in cols:
                        val = row.get(cols[col_name])
                        if val is not None and str(val) != "nan":
                            try:
                                capacity = float(val)
                            except (ValueError, TypeError):
                                pass
                            break

                fuel = None
                for col_name in ["fuel_type", "generation_type", "type",
                                 "fuel", "technology", "resource_type"]:
                    if col_name in cols:
                        val = row.get(cols[col_name])
                        if val and str(val) != "nan":
                            fuel = str(val)
                            break

                status = None
                for col_name in ["status", "queue_status"]:
                    if col_name in cols:
                        val = row.get(cols[col_name])
                        if val and str(val) != "nan":
                            status = str(val)
                            break

                name = None
                for col_name in ["project_name", "name", "facility_name"]:
                    if col_name in cols:
                        val = row.get(cols[col_name])
                        if val and str(val) != "nan":
                            name = str(val)[:200]
                            break

                state = None
                for col_name in ["state", "county_state", "location"]:
                    if col_name in cols:
                        val = row.get(cols[col_name])
                        if val and str(val) != "nan":
                            state = str(val)[:50]
                            break

                rows.append((
                    iso_name, name, None, capacity,
                    fuel, _normalize_fuel(fuel),
                    status, state, None, None, None, None, today,
                ))

            conn.executemany("""
                INSERT OR REPLACE INTO interconnection_queue
                (iso, project_name, developer, capacity_mw, fuel_type, fuel_normalized,
                 status, state, county, queue_date, expected_cod, interconnection_point, fetched_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, rows)
            conn.commit()

            active = [r for r in rows if r[6] and "withdraw" not in (r[6] or "").lower()]
            mw = sum(r[3] or 0 for r in active)
            results[iso_name] = {"projects": len(active), "mw": mw}
            print(f"      {iso_name}: {len(active)} active projects, {mw:,.0f} MW")

        except Exception as e:
            print(f"      {iso_name} FAILED: {e}")
            results[iso_name] = {"projects": 0, "mw": 0, "error": str(e)}

    return results


def fetch_interconnection_queues():
    """Fetch interconnection queue data from US ISOs.

    Uses CAISO direct download + gridstatus for other ISOs.
    Falls back gracefully if gridstatus isn't installed.
    """
    _ensure_tables()
    conn = get_conn()
    today = date.today().isoformat()

    # Check freshness -- only refetch if older than 7 days
    last_fetch = query("SELECT MAX(fetched_date) as d FROM interconnection_queue")
    if last_fetch and last_fetch[0]["d"]:
        days_old = (date.today() - date.fromisoformat(last_fetch[0]["d"])).days
        if days_old < 7:
            print(f"    Queue data is {days_old} days old (threshold: 7). Using cache.")
            return _summarize_queue_data()

    # Clear old data before fresh fetch
    conn.execute("DELETE FROM interconnection_queue")
    conn.commit()

    total_projects = 0
    total_mw = 0

    # 1. CAISO (direct Excel download)
    print("    Fetching CAISO queue (direct download)...")
    caiso = _fetch_caiso_queue(conn, today)
    total_projects += caiso.get("projects", 0)
    total_mw += caiso.get("mw", 0)

    # 2. Other ISOs via gridstatus (if available)
    gs_results = _fetch_gridstatus_queues(conn, today)
    if not gs_results:
        print("    gridstatus not installed -- only CAISO data available.")
        print("    Run: venv/bin/pip install gridstatus  (for PJM, MISO, ERCOT, SPP, NYISO, ISO-NE)")

    for iso_name, stats in gs_results.items():
        total_projects += stats.get("projects", 0)
        total_mw += stats.get("mw", 0)

    print(f"\n    TOTAL: {total_projects} active projects, {total_mw:,.0f} MW")

    # Generate summaries
    summary = _summarize_queue_data()
    _persist_queue_summaries(summary)
    return summary


def _summarize_queue_data():
    """Generate summary statistics from cached queue data."""
    today = date.today().isoformat()

    fuel_summary = query("""
        SELECT fuel_normalized as fuel, COUNT(*) as cnt,
               SUM(capacity_mw) as total_mw, AVG(capacity_mw) as avg_mw
        FROM interconnection_queue
        WHERE status NOT LIKE '%withdraw%' AND status NOT LIKE '%completed%'
        GROUP BY fuel_normalized
        ORDER BY total_mw DESC
    """)

    iso_summary = query("""
        SELECT iso, COUNT(*) as cnt, SUM(capacity_mw) as total_mw
        FROM interconnection_queue
        WHERE status NOT LIKE '%withdraw%' AND status NOT LIKE '%completed%'
        GROUP BY iso
        ORDER BY total_mw DESC
    """)

    cod_dist = query("""
        SELECT SUBSTR(expected_cod, 1, 4) as cod_year,
               fuel_normalized as fuel,
               COUNT(*) as cnt, SUM(capacity_mw) as total_mw
        FROM interconnection_queue
        WHERE expected_cod IS NOT NULL
          AND status NOT LIKE '%withdraw%'
          AND SUBSTR(expected_cod, 1, 4) BETWEEN '2025' AND '2035'
        GROUP BY cod_year, fuel_normalized
        ORDER BY cod_year, total_mw DESC
    """)

    nuclear = query("""
        SELECT iso, project_name, capacity_mw, status, expected_cod, state
        FROM interconnection_queue
        WHERE fuel_normalized = 'Nuclear'
          AND status NOT LIKE '%withdraw%'
        ORDER BY capacity_mw DESC
        LIMIT 30
    """)

    battery = query("""
        SELECT iso, SUM(capacity_mw) as total_mw, COUNT(*) as cnt
        FROM interconnection_queue
        WHERE fuel_normalized = 'Battery Storage'
          AND status NOT LIKE '%withdraw%'
        GROUP BY iso
        ORDER BY total_mw DESC
    """)

    return {
        "date": today,
        "fuel_summary": fuel_summary or [],
        "iso_summary": iso_summary or [],
        "cod_distribution": cod_dist or [],
        "nuclear_pipeline": nuclear or [],
        "battery_by_iso": battery or [],
        "total_projects": sum(r["cnt"] for r in (iso_summary or [])),
        "total_mw": sum(r["total_mw"] or 0 for r in (iso_summary or [])),
    }


def _persist_queue_summaries(summary):
    """Save queue summary stats to DB."""
    conn = get_conn()
    today = date.today().isoformat()

    # Single query to get all ISO×fuel breakdowns with real metrics
    detail_rows = query("""
        SELECT iso, fuel_normalized as fuel,
               SUM(CASE WHEN status NOT LIKE '%withdraw%' THEN 1 ELSE 0 END) as active_cnt,
               SUM(CASE WHEN status NOT LIKE '%withdraw%' THEN capacity_mw ELSE 0 END) as active_mw,
               ROUND(100.0 * SUM(CASE WHEN status LIKE '%withdraw%' THEN 1 ELSE 0 END) / COUNT(*), 1) as withdrawn_pct,
               ROUND(AVG(CASE WHEN queue_date IS NOT NULL AND status NOT LIKE '%withdraw%'
                   THEN julianday('now') - julianday(queue_date) END), 0) as avg_queue_days,
               CAST(AVG(CASE WHEN expected_cod IS NOT NULL AND status NOT LIKE '%withdraw%'
                   THEN CAST(SUBSTR(expected_cod, 1, 4) AS INTEGER) END) AS INTEGER) as median_cod_year
        FROM interconnection_queue
        GROUP BY iso, fuel_normalized
        HAVING active_cnt > 0
    """)

    rows = []
    for r in (detail_rows or []):
        rows.append((
            today, r["iso"], r["fuel"],
            r["active_cnt"], r["active_mw"] or 0,
            r["withdrawn_pct"] or 0, r["avg_queue_days"] or 0, r["median_cod_year"] or 0,
        ))
    if rows:
        conn.executemany("""
            INSERT OR REPLACE INTO interconnection_queue_summary
            (date, iso, fuel_type, active_count, active_mw, withdrawn_pct, avg_queue_days, median_cod_year)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, rows)
        conn.commit()


# ===================================================================
#  2. NREL TECHNOLOGY COST CURVES
# ===================================================================

def fetch_nrel_cost_curves():
    """Fetch NREL ATB cost data from S3.

    Returns dict with current costs, trends, and crossover points.
    """
    _ensure_tables()
    conn = get_conn()

    # Check freshness -- ATB is annual, refetch if >90 days old
    last_fetch = query("SELECT MAX(fetched_date) as d FROM nrel_cost_curves")
    if last_fetch and last_fetch[0]["d"]:
        days_old = (date.today() - date.fromisoformat(last_fetch[0]["d"])).days
        if days_old < 90:
            print(f"    NREL data is {days_old} days old (threshold: 90). Using cache.")
            return _summarize_nrel_data()

    print("    Downloading NREL ATB data...")

    try:
        resp = requests.get(NREL_ATB_URL, timeout=120)
        resp.raise_for_status()
    except Exception as e:
        try:
            fallback_url = "https://oedi-data-lake.s3.amazonaws.com/ATB/electricity/csv/2024/v2.0.0/ATBe.csv"
            print(f"    Primary URL failed ({e}), trying fallback...")
            resp = requests.get(fallback_url, timeout=120)
            resp.raise_for_status()
        except Exception as e2:
            print(f"    NREL download failed: {e2}")
            return _summarize_nrel_data()

    # Parse CSV
    import csv
    import io
    reader = csv.DictReader(io.StringIO(resp.text))

    rows = []
    today = date.today().isoformat()
    seen = set()

    for row in reader:
        tech_raw = row.get("technology", row.get("technology_alias", ""))
        scenario = row.get("scenario", "")
        year_str = row.get("core_metric_variable", row.get("year", ""))
        metric = row.get("core_metric_parameter", row.get("metric", ""))
        value_str = row.get("value", "")
        unit = row.get("units", "")

        # Filter to techs we care about
        tech_lower = tech_raw.lower().replace("-", "").replace(" ", "")
        tech_display = None
        for nrel_key, display in NREL_TECH_MAP.items():
            if nrel_key in tech_lower:
                tech_display = display
                break
        if not tech_display:
            continue

        # Filter scenarios
        scenario_lower = (scenario or "").lower()
        if "moderate" not in scenario_lower and "mid" not in scenario_lower:
            if "advanced" not in scenario_lower and "conservative" not in scenario_lower:
                continue

        # Normalize scenario
        if "advanced" in scenario_lower or "low" in scenario_lower:
            scenario_norm = "Advanced"
        elif "conservative" in scenario_lower or "high" in scenario_lower:
            scenario_norm = "Conservative"
        else:
            scenario_norm = "Moderate"

        # Filter metrics
        metric_upper = (metric or "").upper()
        if not any(m in metric_upper for m in ["LCOE", "CAPEX", "O&M", "CF", "OCC", "CFC"]):
            continue

        # Normalize metric name
        if "LCOE" in metric_upper:
            metric_norm = "LCOE"
        elif "OCC" in metric_upper or "CAPEX" in metric_upper:
            metric_norm = "CAPEX"
        elif "CFC" in metric_upper:
            metric_norm = "CFC"
        elif "FIXED" in metric_upper:
            metric_norm = "Fixed O&M"
        elif "VARIABLE" in metric_upper:
            metric_norm = "Variable O&M"
        elif "CF" in metric_upper:
            metric_norm = "CF"
        else:
            metric_norm = metric

        try:
            year = int(float(year_str))
            value = float(value_str)
        except (ValueError, TypeError):
            continue

        if year < 2020 or year > 2055:
            continue

        key = (tech_display, scenario_norm, year, metric_norm)
        if key in seen:
            continue
        seen.add(key)

        rows.append((tech_display, scenario_norm, year, metric_norm, value, unit, today))

    if rows:
        conn.executemany("""
            INSERT OR REPLACE INTO nrel_cost_curves
            (technology, scenario, year, metric, value, unit, fetched_date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, rows)
        conn.commit()
        print(f"    Loaded {len(rows)} NREL ATB data points")
    else:
        print("    Warning: no matching rows parsed from NREL data")

    return _summarize_nrel_data()


def _summarize_nrel_data():
    """Generate summary from cached NREL data."""
    current_lcoe = query("""
        SELECT technology, scenario, value, unit
        FROM nrel_cost_curves
        WHERE metric = 'LCOE' AND year = 2025 AND scenario = 'Moderate'
        ORDER BY value ASC
    """)

    lcoe_trajectory = query("""
        SELECT technology, year, value
        FROM nrel_cost_curves
        WHERE metric = 'LCOE' AND scenario = 'Moderate'
          AND year IN (2025, 2030, 2035, 2040, 2050)
        ORDER BY technology, year
    """)

    capex_trend = query("""
        SELECT technology, year, value
        FROM nrel_cost_curves
        WHERE metric IN ('CAPEX', 'OCC') AND scenario = 'Moderate'
          AND year IN (2025, 2030, 2040)
        ORDER BY technology, year
    """)

    crossovers = []
    # Find when renewables beat nuclear on LCOE
    nuclear_lcoe = query("""
        SELECT year, MIN(value) as value FROM nrel_cost_curves
        WHERE technology = 'Nuclear' AND metric = 'LCOE' AND scenario = 'Moderate'
        GROUP BY year ORDER BY year
    """)
    for tech in ["Utility Solar", "Onshore Wind", "Solar + Storage"]:
        tech_lcoe = query("""
            SELECT year, MIN(value) as value FROM nrel_cost_curves
            WHERE technology = ? AND metric = 'LCOE' AND scenario = 'Moderate'
            GROUP BY year ORDER BY year
        """, (tech,))
        if nuclear_lcoe and tech_lcoe:
            nuc_dict = {r["year"]: r["value"] for r in nuclear_lcoe}
            for r in tech_lcoe:
                nuc_val = nuc_dict.get(r["year"])
                if nuc_val and r["value"] < nuc_val:
                    crossovers.append({
                        "technology": tech,
                        "crossover_year": r["year"],
                        "tech_lcoe": r["value"],
                        "nuclear_lcoe": nuc_val,
                    })
                    break

    return {
        "current_lcoe": current_lcoe or [],
        "lcoe_trajectory": lcoe_trajectory or [],
        "capex_trend": capex_trend or [],
        "crossovers": crossovers,
        "tech_count": len(set(r["technology"] for r in (current_lcoe or []))),
    }


# ===================================================================
#  3. CONGRESSIONAL + DOE REGULATORY ACTIVITY
# ===================================================================

def fetch_energy_legislation():
    """Fetch energy-related bills from Congress.gov API."""
    if not CONGRESS_API_KEY:
        print("    CONGRESS_API_KEY not set. Skipping legislation scan.")
        print("    Get a free key at https://api.data.gov/signup/")
        return {"error": "no_api_key", "bills": []}

    _ensure_tables()
    conn = get_conn()

    # Check freshness -- weekly refresh
    last_fetch = query("SELECT MAX(fetched_date) as d FROM energy_legislation WHERE congress = ?",
                       (CURRENT_CONGRESS,))
    if last_fetch and last_fetch[0]["d"]:
        days_old = (date.today() - date.fromisoformat(last_fetch[0]["d"])).days
        if days_old < 7:
            print(f"    Legislation data is {days_old} days old (threshold: 7). Using cache.")
            return _summarize_legislation()

    print(f"    Fetching energy bills from {CURRENT_CONGRESS}th Congress...")

    all_bills = []
    energy_keywords = [
        "energy", "power", "electric", "grid", "nuclear", "solar",
        "wind", "gas", "oil", "petroleum", "pipeline", "utility",
        "renewable", "clean", "carbon", "emission", "climate",
        "battery", "storage", "hydrogen", "geothermal", "coal",
        "lng", "ferc", "interconnection", "transmission", "rate",
        "tariff", "critical mineral",
    ]

    for bill_type in ["hr", "s"]:
        offset = 0
        max_pages = 20  # Safety cap: 20 pages × 250 = 5,000 bills max per type
        while offset < max_pages * 250:
            try:
                resp = requests.get(
                    f"{CONGRESS_API_BASE}/bill/{CURRENT_CONGRESS}/{bill_type}",
                    params={
                        "api_key": CONGRESS_API_KEY,
                        "format": "json",
                        "limit": 250,
                        "offset": offset,
                        "sort": "updateDate+desc",
                    },
                    timeout=30,
                )
                resp.raise_for_status()
                data = resp.json()
                bills = data.get("bills", [])
                if not bills:
                    break

                for bill in bills:
                    title = (bill.get("title") or "").lower()
                    if not any(kw in title for kw in energy_keywords):
                        continue

                    bill_id = f"{bill_type}{bill.get('number', '')}-{CURRENT_CONGRESS}"
                    all_bills.append({
                        "bill_id": bill_id,
                        "congress": CURRENT_CONGRESS,
                        "bill_type": bill_type.upper(),
                        "bill_number": bill.get("number"),
                        "title": bill.get("title", ""),
                        "introduced_date": bill.get("introducedDate", ""),
                        "latest_action": bill.get("latestAction", {}).get("text", ""),
                        "latest_action_date": bill.get("latestAction", {}).get("actionDate", ""),
                        "sponsor": "",
                        "cosponsor_count": 0,
                    })

                if len(bills) < 250:
                    break
                offset += 250
                time.sleep(0.5)

            except Exception as e:
                print(f"    Congress API error: {e}")
                break

    print(f"    Found {len(all_bills)} energy-related bills")

    if all_bills:
        scored_bills = _score_legislation(all_bills[:50])

        today = date.today().isoformat()
        rows = []
        for b in scored_bills:
            rows.append((
                b["bill_id"], b["congress"], b["bill_type"], b["bill_number"],
                b["title"], b["introduced_date"],
                b["latest_action"], b["latest_action_date"],
                b.get("sponsor", ""), b.get("cosponsor_count", 0),
                json.dumps(b.get("subjects", [])),
                b.get("status", "introduced"),
                b.get("energy_relevance_score", 0),
                json.dumps(b.get("affected_sectors", [])),
                json.dumps(b.get("affected_tickers", [])),
                b.get("summary", ""),
                today,
            ))

        conn.executemany("""
            INSERT OR REPLACE INTO energy_legislation
            (bill_id, congress, bill_type, bill_number, title, introduced_date,
             latest_action, latest_action_date, sponsor, cosponsor_count,
             subjects, status, energy_relevance_score, affected_sectors,
             affected_tickers, summary, fetched_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, rows)
        conn.commit()
        print(f"    Persisted {len(rows)} scored bills")

    return _summarize_legislation()


def _score_legislation(bills):
    """Use Gemini to score energy bills for market relevance."""
    if not GEMINI_API_KEY or not bills:
        for b in bills:
            score = 0
            title_lower = b["title"].lower()
            if any(kw in title_lower for kw in ["nuclear", "grid", "transmission", "interconnection"]):
                score += 40
            if any(kw in title_lower for kw in ["solar", "wind", "renewable", "clean energy"]):
                score += 30
            if any(kw in title_lower for kw in ["tariff", "lng", "oil", "gas"]):
                score += 25
            if "tax credit" in title_lower or "incentive" in title_lower:
                score += 35
            b["energy_relevance_score"] = min(100, score)
            b["affected_sectors"] = []
            b["affected_tickers"] = []
        return bills

    bill_text = "\n".join(
        f"[{b['bill_id']}] {b['title'][:150]} -- Latest: {b['latest_action'][:80]}"
        for b in bills[:30]
    )

    prompt = f"""Classify these energy bills for stock market impact.

BILLS:
{bill_text}

For each bill_id, respond with JSON array:
[{{"bill_id": "...", "relevance": 0-100, "sectors": ["energy", "utilities", ...], "tickers": ["CEG", "NEE", ...], "summary": "one sentence impact"}}]

Score relevance based on: probability of passing x magnitude of market impact.
High relevance: nuclear/grid expansion bills with bipartisan support.
Low relevance: resolutions, commemorative bills, narrow local bills.

Return ONLY the JSON array, no explanation."""

    try:
        import re
        resp = requests.post(
            f"{GEMINI_BASE}/models/{GEMINI_MODEL}:generateContent",
            headers={"Content-Type": "application/json"},
            params={"key": GEMINI_API_KEY},
            json={
                "contents": [{"role": "user", "parts": [{"text": prompt}]}],
                "generationConfig": {
                    "temperature": 0.1,
                    "maxOutputTokens": 4096,
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            },
            timeout=60,
        )
        resp.raise_for_status()
        data = resp.json()
        parts = data["candidates"][0]["content"]["parts"]
        text = "\n".join(p["text"] for p in parts if "text" in p and not p.get("thought"))

        json_match = re.search(r'\[.*\]', text, re.DOTALL)
        if json_match:
            scored = json.loads(json_match.group())
            score_map = {s["bill_id"]: s for s in scored}

            for b in bills:
                if b["bill_id"] in score_map:
                    s = score_map[b["bill_id"]]
                    b["energy_relevance_score"] = s.get("relevance", 0)
                    b["affected_sectors"] = s.get("sectors", [])
                    b["affected_tickers"] = s.get("tickers", [])
                    b["summary"] = s.get("summary", "")
    except Exception as e:
        print(f"    Gemini classification failed: {e}")
        # Fall back to keyword scoring
        for b in bills:
            if "energy_relevance_score" not in b:
                b["energy_relevance_score"] = 0
                b["affected_sectors"] = []
                b["affected_tickers"] = []

    return bills


def _summarize_legislation():
    """Summarize cached legislation data."""
    high_impact = query("""
        SELECT bill_id, title, energy_relevance_score, affected_sectors,
               affected_tickers, latest_action, latest_action_date, summary
        FROM energy_legislation
        WHERE congress = ? AND energy_relevance_score >= 30
        ORDER BY energy_relevance_score DESC
        LIMIT 20
    """, (CURRENT_CONGRESS,))

    for b in (high_impact or []):
        for field in ["affected_sectors", "affected_tickers"]:
            try:
                b[field] = json.loads(b[field]) if b[field] else []
            except (json.JSONDecodeError, TypeError):
                b[field] = []

    recent = query("""
        SELECT COUNT(*) as cnt, MAX(latest_action_date) as last_date
        FROM energy_legislation
        WHERE congress = ? AND latest_action_date >= date('now', '-30 days')
    """, (CURRENT_CONGRESS,))

    return {
        "high_impact_bills": high_impact or [],
        "total_bills": (query("SELECT COUNT(*) as cnt FROM energy_legislation WHERE congress = ?",
                              (CURRENT_CONGRESS,)) or [{"cnt": 0}])[0]["cnt"],
        "recent_activity_30d": recent[0]["cnt"] if recent else 0,
        "last_action_date": recent[0]["last_date"] if recent else None,
    }


# ===================================================================
#  GENERATE REGULATORY SIGNALS FOR CONVERGENCE
# ===================================================================

def generate_regulatory_signals():
    """Synthesize queue + cost + legislation into actionable signals."""
    _ensure_tables()
    signals = []
    today = date.today().isoformat()
    conn = get_conn()

    # 1. Queue bottleneck signals
    queue_summary = query("""
        SELECT iso, fuel_type, active_count, active_mw
        FROM interconnection_queue_summary
        WHERE date = (SELECT MAX(date) FROM interconnection_queue_summary)
        ORDER BY active_mw DESC
    """)

    if queue_summary:
        solar_mw = sum(r["active_mw"] or 0 for r in queue_summary if r["fuel_type"] == "Solar")
        total_mw = sum(r["active_mw"] or 0 for r in queue_summary)
        if total_mw > 0 and solar_mw / total_mw > 0.5:
            signals.append({
                "date": today, "source": "interconnection_queue",
                "signal_type": "solar_queue_dominance",
                "headline": f"Solar dominates interconnection queue at {solar_mw/total_mw*100:.0f}% of total MW",
                "detail": f"Solar: {solar_mw:,.0f} MW of {total_mw:,.0f} MW total across all ISOs",
                "affected_sectors": json.dumps(["solar", "utilities"]),
                "affected_tickers": json.dumps(QUEUE_AFFECTED_TICKERS.get("Solar", [])),
                "impact_score": 60,
            })

        nuclear_mw = sum(r["active_mw"] or 0 for r in queue_summary if r["fuel_type"] == "Nuclear")
        if nuclear_mw > 5000:
            signals.append({
                "date": today, "source": "interconnection_queue",
                "signal_type": "nuclear_pipeline_growing",
                "headline": f"Nuclear pipeline: {nuclear_mw:,.0f} MW in interconnection queues",
                "detail": f"Nuclear renaissance signal -- {nuclear_mw:,.0f} MW seeking grid connection",
                "affected_sectors": json.dumps(["nuclear", "utilities"]),
                "affected_tickers": json.dumps(QUEUE_AFFECTED_TICKERS.get("Nuclear", [])),
                "impact_score": 75,
            })

        battery_mw = sum(r["active_mw"] or 0 for r in queue_summary if r["fuel_type"] == "Battery Storage")
        if battery_mw > 50000:
            signals.append({
                "date": today, "source": "interconnection_queue",
                "signal_type": "battery_storage_surge",
                "headline": f"Battery storage: {battery_mw:,.0f} MW in queue -- grid flexibility play",
                "detail": "Massive storage buildout signals grid modernization acceleration",
                "affected_sectors": json.dumps(["storage", "utilities"]),
                "affected_tickers": json.dumps(QUEUE_AFFECTED_TICKERS.get("Battery Storage", [])),
                "impact_score": 65,
            })

    # 2. Cost curve signals
    crossovers = query("""
        SELECT technology, year, value FROM nrel_cost_curves
        WHERE metric = 'LCOE' AND scenario = 'Moderate' AND year = 2025
        ORDER BY value ASC
    """)
    if crossovers and len(crossovers) >= 2:
        cheapest = crossovers[0]
        signals.append({
            "date": today, "source": "nrel_atb",
            "signal_type": "cost_leadership",
            "headline": f"{cheapest['technology']} is cheapest new-build at ${cheapest['value']:.0f}/MWh LCOE",
            "detail": f"NREL ATB 2025: {cheapest['technology']} leads on unsubsidized cost",
            "affected_sectors": json.dumps(["energy", "utilities"]),
            "affected_tickers": json.dumps([]),
            "impact_score": 55,
        })

    # 3. Legislative signals
    high_bills = query("""
        SELECT bill_id, title, energy_relevance_score, affected_tickers
        FROM energy_legislation
        WHERE congress = ? AND energy_relevance_score >= 60
        ORDER BY energy_relevance_score DESC
        LIMIT 5
    """, (CURRENT_CONGRESS,))

    for bill in (high_bills or []):
        tickers = []
        try:
            tickers = json.loads(bill["affected_tickers"]) if bill["affected_tickers"] else []
        except (json.JSONDecodeError, TypeError):
            pass
        signals.append({
            "date": today, "source": "congress",
            "signal_type": "high_impact_legislation",
            "headline": f"[{bill['bill_id']}] {bill['title'][:120]}",
            "detail": f"Relevance: {bill['energy_relevance_score']}/100",
            "affected_sectors": json.dumps(["energy", "utilities"]),
            "affected_tickers": json.dumps(tickers),
            "impact_score": bill["energy_relevance_score"],
        })

    # Persist
    if signals:
        rows = [(s["date"], s["source"], s["signal_type"], s["headline"],
                 s["detail"], s["affected_sectors"], s["affected_tickers"],
                 s["impact_score"]) for s in signals]
        conn.executemany("""
            INSERT OR REPLACE INTO energy_regulatory_signals
            (date, source, signal_type, headline, detail, affected_sectors,
             affected_tickers, impact_score)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, rows)
        conn.commit()
        print(f"\n    Generated {len(signals)} regulatory signals")

    return signals


# ===================================================================
#  CLI ENTRY POINT
# ===================================================================

def run():
    parser = argparse.ArgumentParser(description="Energy Infrastructure Intelligence")
    parser.add_argument("--queues", action="store_true", help="Fetch interconnection queue data")
    parser.add_argument("--costs", action="store_true", help="Fetch NREL cost curves")
    parser.add_argument("--regulatory", action="store_true", help="Fetch Congressional + DOE activity")
    parser.add_argument("--all", action="store_true", help="Run all three data sources")
    parser.add_argument("--signals", action="store_true", help="Generate regulatory signals only (from cache)")
    args = parser.parse_args()

    if not any([args.queues, args.costs, args.regulatory, args.all, args.signals]):
        args.all = True

    init_db()
    _ensure_tables()

    print(f"\n{'=' * 60}")
    print(f"  ENERGY INFRASTRUCTURE INTELLIGENCE")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'=' * 60}")

    if args.all or args.queues:
        print("\n  [1/3] Interconnection Queues")
        print("  " + "-" * 40)
        queue_data = fetch_interconnection_queues()
        if "error" not in queue_data:
            print(f"    Total: {queue_data.get('total_projects', 0):,} projects, "
                  f"{queue_data.get('total_mw', 0):,.0f} MW")
            if queue_data.get("nuclear_pipeline"):
                print(f"    Nuclear pipeline: {len(queue_data['nuclear_pipeline'])} projects")

    if args.all or args.costs:
        print("\n  [2/3] NREL Technology Cost Curves")
        print("  " + "-" * 40)
        cost_data = fetch_nrel_cost_curves()
        if cost_data.get("current_lcoe"):
            print("    Current LCOE ($/MWh, Moderate scenario):")
            for r in cost_data["current_lcoe"][:8]:
                print(f"      {r['technology']:25s} ${r['value']:.0f}")
        if cost_data.get("crossovers"):
            print("    Cost crossovers vs Nuclear:")
            for c in cost_data["crossovers"]:
                print(f"      {c['technology']} beats Nuclear by {c['crossover_year']}")

    if args.all or args.regulatory:
        print("\n  [3/3] Congressional Activity")
        print("  " + "-" * 40)
        leg_data = fetch_energy_legislation()
        print(f"    Total energy bills: {leg_data.get('total_bills', 0)}")
        print(f"    High-impact (score>=30): {len(leg_data.get('high_impact_bills', []))}")
        if leg_data.get("high_impact_bills"):
            print("    Top bills:")
            for b in leg_data["high_impact_bills"][:5]:
                print(f"      [{b['energy_relevance_score']:.0f}] {b['title'][:100]}")

    if args.all or args.signals:
        print("\n  Generating Regulatory Signals...")
        print("  " + "-" * 40)
        signals = generate_regulatory_signals()
        for s in signals:
            print(f"    [{s['impact_score']:.0f}] {s['headline'][:100]}")

    print(f"\n{'=' * 60}")
    print(f"  COMPLETE")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    run()
